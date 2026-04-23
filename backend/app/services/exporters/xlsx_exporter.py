"""XLSX cost + material schedule — openpyxl.

Four sheets:
  - Summary          : project meta
  - Materials        : flattened material spec
  - Cost Breakdown   : line items + totals
  - MEP Schedule     : HVAC / electrical / plumbing key values
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


INK = "FF1F1D1A"
INK_SOFT = "FF4A463F"
PAPER = "FFF7F2EA"
PAPER_DEEP = "FFECE5D8"

_HEADER_FILL = PatternFill(start_color=INK, end_color=INK, fill_type="solid")
_HEADER_FONT = Font(name="Calibri", size=10, bold=True, color=PAPER)
_BODY_FONT = Font(name="Calibri", size=10, color=INK)
_THIN = Side(style="thin", color=INK_SOFT)
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _fmt_range(v) -> str:
    if isinstance(v, (list, tuple)) and len(v) == 2:
        return f"{v[0]} – {v[1]}"
    if v is None:
        return "—"
    return str(v)


def _flat(v) -> str:
    if isinstance(v, dict):
        return ", ".join(f"{k}: {_flat(x)}" for k, x in v.items())
    if isinstance(v, list):
        return "; ".join(_flat(x) for x in v)
    if isinstance(v, tuple):
        return _fmt_range(list(v))
    return "" if v is None else str(v)


def _write_header(ws, row: int, header: list[str]) -> None:
    for i, h in enumerate(header, start=1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = _BORDER


def _write_rows(ws, start_row: int, rows: list[list]) -> None:
    for r, row in enumerate(rows, start=start_row):
        for c, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value="" if value is None else value)
            cell.font = _BODY_FONT
            cell.border = _BORDER
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)


def _autosize(ws, max_width: int = 48) -> None:
    for col_idx, col in enumerate(ws.columns, start=1):
        letter = get_column_letter(col_idx)
        longest = 12
        for cell in col:
            try:
                v = str(cell.value) if cell.value is not None else ""
                longest = max(longest, min(len(v) + 2, max_width))
            except Exception:
                pass
        ws.column_dimensions[letter].width = longest


def export(spec: dict, graph: dict) -> dict:
    wb = Workbook()
    meta = spec["meta"]

    # Sheet 1 — Summary.
    ws = wb.active
    ws.title = "Summary"
    _write_header(ws, 1, ["Field", "Value"])
    summary_rows = [
        ["Project", meta["project_name"]],
        ["Generated", meta["generated_at"]],
        ["Room type", meta.get("room_type", "—")],
        ["Theme", meta.get("theme", "—")],
        ["Length (m)", meta["dimensions_m"].get("length")],
        ["Width (m)", meta["dimensions_m"].get("width")],
        ["Height (m)", meta["dimensions_m"].get("height")],
        ["Objects", spec.get("objects_count")],
    ]
    _write_rows(ws, 2, summary_rows)
    _autosize(ws)

    # Sheet 2 — Materials.
    ws2 = wb.create_sheet("Materials")
    header = ["Bucket", "Name", "Grade", "Finish", "Color", "Supplier", "Lead (wk)", "Cost INR", "Unit"]
    _write_header(ws2, 1, header)
    rows: list[list] = []
    mat = spec["material"]
    buckets = [
        ("Primary", mat["primary_structure"]),
        ("Secondary", mat["secondary_materials"]),
        ("Upholstery", mat["upholstery"]),
        ("Hardware", mat["hardware"]),
        ("Finishing", mat["finishing"]),
    ]
    for label, bucket_rows in buckets:
        for r in bucket_rows:
            rows.append([
                label,
                r.get("name", "—"),
                r.get("grade", "—"),
                r.get("finish", "—"),
                r.get("color", "—"),
                r.get("supplier", "—"),
                _fmt_range(r.get("lead_time_weeks")),
                _fmt_range(r.get("cost_inr")),
                r.get("unit", "—"),
            ])
    _write_rows(ws2, 2, rows)
    _autosize(ws2)

    # Sheet 3 — Cost.
    ws3 = wb.create_sheet("Cost Breakdown")
    c = spec["cost"]
    totals = c.get("totals", {})
    _write_header(ws3, 1, ["Category", "Item", "Qty", "Unit", "Rate", "Low", "High"])
    cost_rows: list[list] = []
    for li in c.get("line_items") or []:
        cost_rows.append([
            li.get("category", "—"),
            li.get("itemName") or li.get("item_name") or li.get("name", "—"),
            li.get("quantity", "—"),
            li.get("unit", "—"),
            _fmt_range(li.get("unitRate") or li.get("unit_rate")),
            _fmt_range(li.get("totalLow") or li.get("total_low")),
            _fmt_range(li.get("totalHigh") or li.get("total_high")),
        ])
    _write_rows(ws3, 2, cost_rows)
    # Totals block.
    total_row_start = len(cost_rows) + 3
    _write_rows(ws3, total_row_start, [
        ["TOTAL LOW", "", "", "", "", totals.get("low", "—"), ""],
        ["TOTAL BASE", "", "", "", "", totals.get("base", "—"), ""],
        ["TOTAL HIGH", "", "", "", "", "", totals.get("high", "—")],
    ])
    _autosize(ws3)

    # Sheet 4 — MEP.
    ws4 = wb.create_sheet("MEP Schedule")
    _write_header(ws4, 1, ["System", "Parameter", "Value"])
    mep_rows: list[list] = []
    for system, block in spec["mep"].items():
        for k, v in block.items():
            mep_rows.append([system.upper(), k.replace("_", " ").title(), _flat(v)])
    _write_rows(ws4, 2, mep_rows)
    _autosize(ws4)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return {
        "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "filename": f"{_safe_name(meta['project_name'])}-schedule.xlsx",
        "bytes": buffer.getvalue(),
    }


def _safe_name(name: str) -> str:
    return "".join(c if c.isalnum() or c in "-_" else "-" for c in name).strip("-") or "project"
