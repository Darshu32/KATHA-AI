"""XLSX importer — sheet-by-sheet header + sample-row extraction (openpyxl)."""

from __future__ import annotations

import io
from typing import Any

try:
    from openpyxl import load_workbook
except Exception:  # noqa: BLE001
    load_workbook = None

from app.services.importers.csv_importer import _classify


def parse(filename: str, payload: bytes) -> dict[str, Any]:
    if load_workbook is None:
        return {
            "format": "xlsx",
            "filename": filename,
            "size_bytes": len(payload),
            "summary": "openpyxl not installed.",
            "extracted": {},
            "warnings": ["openpyxl is not available — install to parse .xlsx files."],
        }
    try:
        wb = load_workbook(io.BytesIO(payload), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        return {
            "format": "xlsx",
            "filename": filename,
            "size_bytes": len(payload),
            "summary": "Could not parse XLSX.",
            "extracted": {},
            "warnings": [f"openpyxl load failed: {exc}"],
        }

    sheets_out: list[dict[str, Any]] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            sheets_out.append({"name": sheet_name, "row_count": 0, "headers": [], "sample_rows": []})
            continue
        headers = [str(c) if c is not None else "" for c in header_row]
        body: list[list[Any]] = []
        for i, row in enumerate(rows_iter):
            body.append(list(row))
            if i >= 50:
                break
        kinds = _classify(headers)
        sheets_out.append({
            "name": sheet_name,
            "headers": headers,
            "row_count": ws.max_row - 1 if ws.max_row else len(body),
            "kinds_detected": kinds,
            "sample_rows": [dict(zip(headers, r)) for r in body[:10]],
        })
    return {
        "format": "xlsx",
        "filename": filename,
        "size_bytes": len(payload),
        "summary": (
            f"XLSX: {len(sheets_out)} sheet(s); "
            f"row counts {sum(s.get('row_count') or 0 for s in sheets_out)}."
        ),
        "extracted": {"sheets": sheets_out},
        "warnings": [],
    }
